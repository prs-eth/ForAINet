import os
import numpy as np
from scipy import stats
from utils.ply import read_ply, write_ply
from utils.utils import output_DTM_as_pc, DTM_generation, cal_DBH_and_centerP, DTM_accuracy, hdbscan_filtering, preprocess_point_cloud,alpha_shape_volume,plot_alpha_shape, compute_volume_with_voxelization, compute_volume_with_convex_hull
from os.path import exists, join
import matplotlib.pyplot as plt
import pyransac3d as pyrsc
from matplotlib.patches import Circle
import imageio
from skimage import color
from skimage import io
from RANSAC.RANSAC.RANSACCircle_2 import run
from plyfile import PlyData, PlyElement
from RANSAC.smallest_enclosing_circle import welzl
import math
from RANSAC.Common.Point import Point
from typing import List
from RANSAC.RANSAC.MatplotUtil import create_list_from_points
import pandas as pd
from scipy.interpolate import RectBivariateSpline, interp2d, bisplrep, bisplev, griddata
from scipy.spatial import ConvexHull
from openpyxl import Workbook
from sklearn.metrics import confusion_matrix
from tree_metrics.rmse import rmse
fold= ['/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/CULS/CULS_plot_2_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/NIBIO/NIBIO_plot_1_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/TUWIEN/TUWIEN_test_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/NIBIO/NIBIO_plot_17_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/NIBIO/NIBIO_plot_18_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/NIBIO/NIBIO_plot_22_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/NIBIO/NIBIO_plot_23_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/NIBIO/NIBIO_plot_5_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/RMIT/RMIT_test_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/SCION/SCION_plot_31_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/SCION/SCION_plot_61_annotated_test.ply']
#Input files directory
file_path = '/scratch2/OutdoorPanopticSeg_V2/outputs/best_treemix2/eval/2023-09-21_20-23-39/'
all_files = os.listdir(file_path)
matching_files = [filename for filename in all_files if 'Semantic_results_forEval' in filename]
#output files directory
directory = file_path+'para_cal_imgs'
if not os.path.exists(directory):
    os.makedirs(directory)

#predefined parameters
BINSIZE = 0.5 #for dtm rasterization
height_of_trunkDiameter = 1.3
#th_height_of_tree = 5
thre_dbh = 0.1  #thresholf for dominant and non-dominant tree

#Iterate over all plots
for f_idx, filename in enumerate(matching_files):
    #predicted semantic segmentation file path
    pred_class_label_filename = file_path + filename
    #predicted instance segmentation file path
    pred_ins_label_filename = file_path +filename.replace('Semantic_results_forEval', 'Instance_Results_forEval')

    new_dir_name = "plot" + pred_class_label_filename[-6:-4]
    # Construct the full path to the new directory for new plot
    current_plot = os.path.join(directory, new_dir_name)
    # Check if the directory exists, and create it if it doesn't
    if os.path.exists(current_plot):
        continue
    if not os.path.exists(current_plot):
        os.mkdir(current_plot)

    #read files
    data_class = PlyData.read(pred_class_label_filename) 
    data_ins = PlyData.read(pred_ins_label_filename)
    pred_ins_complete = data_ins['vertex']['preds']
    pred_sem_complete = data_class['vertex']['preds']
    gt_ins_complete = data_ins['vertex']['gt'] -1
    gt_sem_complete = data_class['vertex']['gt']

    #idx for ground points
    idx_gt_ground = (gt_sem_complete==1)
    idx_pre_ground = (pred_sem_complete==1)

    ##########################Attributes for plot-level###########################
    #DTM generation from ground points
    #DTM of gt
    dtm_gt = DTM_generation(data_ins['vertex']['x'][idx_gt_ground], data_ins['vertex']['y'][idx_gt_ground], data_ins['vertex']['z'][idx_gt_ground], BINSIZE)
    #DTM of prediction
    dtm_pre = DTM_generation(data_ins['vertex']['x'][idx_pre_ground], data_ins['vertex']['y'][idx_pre_ground], data_ins['vertex']['z'][idx_pre_ground], BINSIZE)
    #output DTM as point cloud
    pc_out_dest = current_plot+'/floor_gt_pynn'+pred_class_label_filename[-6:-4]+'.ply'
    output_DTM_as_pc(dtm_gt, pc_out_dest)
    pc_out_dest = current_plot+'/floor_pre_pynn'+pred_class_label_filename[-6:-4]+'.ply'
    output_DTM_as_pc(dtm_pre, pc_out_dest)
    
    #The accuracy of the DTM
    dtm_coverage, dtm_RMSE = DTM_accuracy(data_ins['vertex'], BINSIZE, idx_gt_ground, idx_pre_ground)
    
    #Stem density for classify complexity of the plot
    #idx for vegetation points
    idx_gt_tree = (gt_sem_complete==2)|(gt_sem_complete==3)|(gt_sem_complete==4)
    idx_pre_tree = (pred_sem_complete==2)|(pred_sem_complete==3)|(pred_sem_complete==4)

    #GT instances for vegetation
    x_veg_gt = data_ins['vertex']['x'][idx_gt_tree]
    y_veg_gt = data_ins['vertex']['y'][idx_gt_tree]
    z_veg_gt = data_ins['vertex']['z'][idx_gt_tree]
    veg_ins_gt = gt_ins_complete[idx_gt_tree]
    veg_sem_gt = gt_sem_complete[idx_gt_tree]
    cor_veg_ins_pre = -1*np.ones_like(veg_ins_gt)
    veg_height_gt = np.zeros_like(veg_ins_gt, dtype=float)

    #predicted instances for vegetation
    x_veg_pre = data_ins['vertex']['x'][idx_pre_tree]
    y_veg_pre = data_ins['vertex']['y'][idx_pre_tree]
    z_veg_pre = data_ins['vertex']['z'][idx_pre_tree]
    veg_ins_pre = pred_ins_complete[idx_pre_tree]
    veg_sem_pre = pred_sem_complete[idx_pre_tree]
    veg_height_pre = np.zeros_like(veg_ins_pre, dtype=float)

    
    #tree instances in prediction set
    # initiate plot
    fig = plt.figure()
    #print(fig.dpi)  #dpi=100
    un = np.unique(veg_ins_pre)
    pts_in_pred = []
    pts_in_pred_tree = []
    #spline of ground
    interp_spline = bisplrep(dtm_pre[:,0], dtm_pre[:,1], dtm_pre[:,2])
    ##########################Attributes for tree-level###########################
    #loop for each predicted tree instance
    for ig, g in enumerate(un):
        fig.clf() # Clear old figure
        ax = fig.add_subplot(2, 4, 1, projection='3d')
        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        ax.set_zlabel('Z')
        if g == -1:
            continue
        tmp = np.where(veg_ins_pre == g)
        x_tmp = x_veg_pre[tmp]
        y_tmp = y_veg_pre[tmp]
        z_tmp = np.copy(z_veg_pre[tmp])
        
        #get ground height based on DTM
        x_center = np.mean(x_tmp)
        y_center = np.mean(y_tmp)
        ground_z = bisplev(x_center, y_center, interp_spline)
        #ground_z = griddata((dtm_pre[:,0], dtm_pre[:,1]), dtm_pre[:,2], (x_tmp, y_tmp), method='linear')
        #z value above ground
        z_tmp = z_tmp-ground_z
        
        tmp_tree = np.where(veg_ins_pre == g)
        tmp_com = np.where(pred_ins_complete == g)
        pts_in_pred += [tmp_com[0]]
        pts_in_pred_tree += [tmp_tree[0]]
        
        #show current tree instance to be calculated
        ax.scatter(x_tmp, y_tmp, z_tmp, marker='o', s=0.2)
        plt.pause(1)
        
        #hdbscan filtering
        xyz_tmp = np.concatenate((np.concatenate((np.expand_dims(x_tmp,axis=1),np.expand_dims(y_tmp,axis=1)),axis=1), np.expand_dims(z_tmp,axis=1)), axis=1)
        filtered_points, con_4 = hdbscan_filtering(xyz_tmp) #, distance_threshold=0.3)
        
        # Set the minimum points threshold
        min_points_threshold = 10
        im_path = current_plot+'/'+ 'preTree'+str(g)+ '_fittingPointsProj.png'
        
        if np.size(con_4)<min_points_threshold:
            fig1_path = im_path.replace('fittingPointsProj', '')
            plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=0.5, hspace=0.5)
            fig.savefig(fig1_path, bbox_inches='tight', pad_inches=0)
            continue
        veg_height_pre[tmp] = np.max(filtered_points[:,2])
        #veg_xloc_pre[tmp] = np.mean(filtered_points[:,0])
        #veg_yloc_pre[tmp] = np.mean(filtered_points[:,1])
         
        ax = fig.add_subplot(2, 4, 2, projection='3d')
        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        ax.set_zlabel('Z')
        #show tree instance after filtering
        ax.scatter(filtered_points[:,0], filtered_points[:,1], filtered_points[:,2], marker='o', s=0.2)
        plt.pause(1)
        
        fig1_path = im_path.replace('fittingPointsProj', '')
        
        plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=0.5, hspace=0.5)
        fig.savefig(fig1_path, bbox_inches='tight', pad_inches=0)
    plt.close()  
    #tree instances in GT set
    fig = plt.figure()
    #print(fig.dpi)  #dpi=100
    un = np.unique(veg_ins_gt)
    pts_in_gt = []
    pts_in_gt_tree = []
    interp_spline = bisplrep(dtm_gt[:,0], dtm_gt[:,1], dtm_gt[:,2])
    for ig, g in enumerate(un):
        fig.clf() # Clear figure
        ax = fig.add_subplot(2, 4, 1, projection='3d')
        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        ax.set_zlabel('Z')
        if g == -1:
            continue
        tmp = np.where(veg_ins_gt == g)
        x_tmp = x_veg_gt[tmp]
        y_tmp = y_veg_gt[tmp]
        z_tmp = np.copy(z_veg_gt[tmp])
        
        #get ground height based on DTM
        x_center = np.mean(x_tmp)
        y_center = np.mean(y_tmp)
        ground_z = bisplev(x_center, y_center, interp_spline)
        #ground_z = griddata((dtm_pre[:,0], dtm_pre[:,1]), dtm_pre[:,2], (x_tmp, y_tmp), method='linear')
        #z value above ground
        z_tmp = z_tmp-ground_z
        
        tmp_tree = np.where(veg_ins_gt == g)
        tmp_com = np.where(gt_ins_complete == g) 
        pts_in_gt += [tmp_com[0]]
        pts_in_gt_tree += [tmp_tree[0]]
        
        #show current instance to be calculated
        ax.scatter(x_tmp, y_tmp, z_tmp, marker='o', s=0.2)
        plt.pause(1)
        
        #hdbscan filtering
        xyz_tmp = np.concatenate((np.concatenate((np.expand_dims(x_tmp,axis=1),np.expand_dims(y_tmp,axis=1)),axis=1), np.expand_dims(z_tmp,axis=1)), axis=1)
        filtered_points, con_4 = hdbscan_filtering(xyz_tmp)
        
        im_path = current_plot+'/'+ 'gtTree'+str(g)+ '_fittingPointsProj.png'
        if np.size(con_4)<min_points_threshold:
            fig1_path = im_path.replace('fittingPointsProj', '')
            plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=0.5, hspace=0.5)
            fig.savefig(fig1_path, bbox_inches='tight', pad_inches=0)
            continue
        veg_height_gt[tmp] = np.max(filtered_points[:,2])  #np.max(z_tmp)
        
        #xyz_tmp = np.concatenate((np.concatenate((np.expand_dims(x_tmp,axis=1),np.expand_dims(y_tmp,axis=1)),axis=1), np.expand_dims(z_tmp,axis=1)), axis=1)
        
        ax = fig.add_subplot(2, 4, 2, projection='3d')
        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        ax.set_zlabel('Z')
        #show tree instance after filtering
        ax.scatter(filtered_points[:,0], filtered_points[:,1], filtered_points[:,2], marker='o', s=0.2)
        plt.pause(1)
        
        fig1_path = im_path.replace('fittingPointsProj', '')
        
        plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=0.5, hspace=0.5)
        fig.savefig(fig1_path, bbox_inches='tight', pad_inches=0)
        
    plt.close()   
    
    ####################### classify dominant and non-dominant tree#######################
    pts_in_gt_dominant = []
    pts_in_gt_small = []
    pts_in_gt_tree_dominant = []
    pts_in_gt_tree_small = []
    threshold_h = np.max(veg_height_gt)/3
    for ig, ins_gt in enumerate(pts_in_gt):
        cor_gt = pts_in_gt_tree[ig]
        height_temp = veg_height_gt[cor_gt][0]
        #if height_temp > threshold_h:
        if height_temp <= threshold_h:
            pts_in_gt_dominant += [ins_gt]
            pts_in_gt_tree_dominant += [cor_gt]
        else:
            pts_in_gt_small += [ins_gt]
            pts_in_gt_tree_small += [cor_gt]
            
    pts_in_pred_dominant = []
    pts_in_pred_small = []
    pts_in_pred_tree_dominant = []
    pts_in_pred_tree_small = []
    for ig, ins_pre in enumerate(pts_in_pred):
        cor_pre = pts_in_pred_tree[ig]
        height_temp = veg_height_pre[cor_pre][0]
        #if height_temp > threshold_h:
        if height_temp <= threshold_h:
            pts_in_pred_dominant += [ins_pre]
            pts_in_pred_tree_dominant += [cor_pre]
        else:
            pts_in_pred_small += [ins_pre]
            pts_in_pred_tree_small += [cor_pre]
    ###################### Evaluate tree measurements with regard to GT dominant#######################
    cor_gt = []
    cor_pre = []
    gt_H = []
    error_H = []
    total_gt_ins_dominant = len(pts_in_gt_dominant)  #No_reference_trees
    matched_gt_ins_dominant = 0  #No_correct_detected_trees
    No_detected_trees_dominant = len(pts_in_pred_dominant)
    at=0.5
    sum_IoU = 0
    # loop for each GT instance
    for ig, ins_gt in enumerate(pts_in_gt_dominant):
        ovmax = 0
        # find the best matched prediction
        for ip, ins_pred in enumerate(pts_in_pred_dominant):
            union = len(list(set(ins_pred).union(set(ins_gt)))) 
            intersect = len(list(set(ins_pred).intersection(set(ins_gt)))) 
            iou = float(intersect) / union 

            if iou > ovmax:
                ovmax = iou
                ipmax = ip

        if ovmax >= at:
            matched_gt_ins_dominant+=1
            cor_gt = pts_in_gt_tree_dominant[ig]
            cor_pre = pts_in_pred_tree_dominant[ipmax]
            cor_veg_ins_pre[cor_gt] = pred_ins_complete[pts_in_pred_dominant[ipmax][0]]
            error_H += [veg_height_pre[cor_pre][0]-veg_height_gt[cor_gt][0]]
            #[np.square(veg_height_pre[cor_pre][0]-veg_height_gt[cor_gt][0])]
            gt_H += [veg_height_gt[cor_gt][0]]
        sum_IoU += ovmax
    
    mean_IoU_dominant = (sum_IoU/total_gt_ins_dominant)*100
    No_omitted_trees_dominant = total_gt_ins_dominant-matched_gt_ins_dominant
    No_wrong_detected_trees_dominant = No_detected_trees_dominant-matched_gt_ins_dominant
    dr_dominant = (No_detected_trees_dominant/total_gt_ins_dominant)*100 
    da_dominant = (matched_gt_ins_dominant/total_gt_ins_dominant)*100  #completeness rate/tree detection accuracy
    omission_error_dominant = 100-da_dominant
    commission_error_dominant = (No_wrong_detected_trees_dominant/No_detected_trees_dominant)*100
    F1_score_dominant = (2*(100-commission_error_dominant)*da_dominant)/(100-commission_error_dominant+da_dominant)
    
            
    # Evaluate tree measurements with regard to GT
    cor_gt = []
    cor_pre = []
    gt_H = []
    error_H = []
    total_gt_ins = len(pts_in_gt)  #No_reference_trees
    matched_gt_ins = 0  #No_correct_detected_trees
    No_detected_trees = len(pts_in_pred)
    at=0.5
    sum_IoU = 0
    # loop for each GT instance
    for ig, ins_gt in enumerate(pts_in_gt):
        ovmax = 0
        # find the best matched prediction
        for ip, ins_pred in enumerate(pts_in_pred):
            union = len(list(set(ins_pred).union(set(ins_gt)))) 
            intersect = len(list(set(ins_pred).intersection(set(ins_gt)))) 
            iou = float(intersect) / union 

            if iou > ovmax:
                ovmax = iou
                ipmax = ip

        if ovmax >= at:
            matched_gt_ins+=1
            cor_gt = pts_in_gt_tree[ig]
            cor_pre = pts_in_pred_tree[ipmax]
            cor_veg_ins_pre[cor_gt] = pred_ins_complete[pts_in_pred[ipmax][0]]
            error_H += [veg_height_pre[cor_pre][0]-veg_height_gt[cor_gt][0]]
            #[np.square(veg_height_pre[cor_pre][0]-veg_height_gt[cor_gt][0])]
            gt_H += [veg_height_gt[cor_gt][0]]
        sum_IoU += ovmax
    
    mean_IoU = (sum_IoU/total_gt_ins)*100
    No_omitted_trees = total_gt_ins-matched_gt_ins
    No_wrong_detected_trees = No_detected_trees-matched_gt_ins
    dr = (No_detected_trees/total_gt_ins)*100 
    da = (matched_gt_ins/total_gt_ins)*100  #completeness rate/tree detection accuracy
    omission_error = 100-da
    commission_error = (No_wrong_detected_trees/No_detected_trees)*100
    F1_score = (2*(100-commission_error)*da)/(100-commission_error+da)
    
    #RMSE, Root Mean Squared Error
    gt_h_hat = np.mean(gt_H)
    RMSE_H = np.sqrt(np.mean(np.square(error_H)))
    RMSE_H_100 = RMSE_H/gt_h_hat
    bias_H = np.mean(error_H)
    print('true positive={}%, {} out of {} instances'.format((matched_gt_ins/total_gt_ins)*100, matched_gt_ins, total_gt_ins))
    print('RMSE for vegetation height: {}'.format(RMSE_H))

    # Project the point of trees onto the xy plane
    projected_points = np.concatenate((np.expand_dims(x_veg_gt,axis=1),np.expand_dims(y_veg_gt,axis=1)),axis=1)
    # Calculate the convex hull of the projected points
    hull = ConvexHull(projected_points)
    # Calculate the area of the convex hull polygon
    convex_area = hull.area
    stem_density = (total_gt_ins/convex_area*10000)  #square meters to hectares, 1 hectare (ha) = 10,000 square meters (m²)
    
    
    dominant_array = np.zeros_like(veg_height_gt)
    dominant_array[veg_height_gt > threshold_h] = 1
    output_path = join(current_plot, 'vegetable_gt_total.ply')
    write_ply(output_path,
          [x_veg_gt, y_veg_gt, z_veg_gt, veg_ins_gt, veg_height_gt, cor_veg_ins_pre, dominant_array],
          ['x', 'y', 'z', 'ins_label', 'height_gt', 'pre_label', 'dominant'])
    output_path2 = join(current_plot, 'vegetable_pre_total.ply')
    dominant_array = np.zeros_like(veg_height_pre)
    dominant_array[veg_height_pre > threshold_h] = 1
    write_ply(output_path2,
          [x_veg_pre, y_veg_pre, z_veg_pre, veg_ins_pre, veg_height_pre, dominant_array],
          ['x', 'y', 'z', 'ins_label', 'height_pre', 'dominant'])

    veg_ins_gt_u, indices = np.unique(veg_ins_gt, return_index=True)
    cor_pre_ins = [] #-1*np.ones_like(indices)
    cor_veg_height_pre = []

    for i in cor_veg_ins_pre[indices]:
        if i==-1:
            #idx_for_pre.append(0)
            cor_pre_ins.append(-1)
            cor_veg_height_pre.append(-1)
        else:
            inn=np.where(veg_ins_pre==i)[0][0] 
            cor_pre_ins.append(veg_ins_pre[inn])
            cor_veg_height_pre.append(veg_height_pre[inn])
    cor_pre_ins = np.array(cor_pre_ins)
    cor_veg_height_pre = np.array(cor_veg_height_pre)
    
    c_ind = np.where(veg_height_gt[indices]!=0)
    
    output = np.hstack((veg_ins_gt[indices[c_ind]].reshape(-1,1),cor_pre_ins[c_ind].reshape(-1,1),
            veg_height_gt[indices[c_ind]].reshape(-1,1), cor_veg_height_pre[c_ind].reshape(-1,1)
            ))
    df1 = pd.DataFrame(output,
                    dtype=np.float32,
                    columns=['gt label', 'pre label',
                                'gt height', 'pre height'])
    # Combine the additional metrics into a new DataFrame
    additional_metrics_data = {
        'dtm_coverage (%)' : dtm_coverage,
        'dtm_RMSE' : dtm_RMSE,
        'No. reference trees': total_gt_ins,
        'No. detected trees': No_detected_trees,
        'DR (%)': dr,
        'No. correctly detected trees': matched_gt_ins,
        'DA (%)': da,
        'No. omitted trees': No_omitted_trees,
        'omission_error (%)': omission_error,
        'No. wrong detections': No_wrong_detected_trees,
        'commission_error (%)': commission_error,
        'F1_score (%)': F1_score,
        'mean_IoU (coverage)': mean_IoU,
        'stem_density': stem_density,
        'RMSE for tree height': RMSE_H,
        'RMSE% for tree height': RMSE_H_100,
        'bias for tree height': bias_H,
    }
    #Convert the dictionary to a list of dictionaries
    additional_metrics_list = [additional_metrics_data]
    additional_metrics_df = pd.DataFrame(additional_metrics_list)
    
    # Combine the additional metrics into a new DataFrame
    additional_metrics_data2 = {
        'No. reference trees': total_gt_ins_dominant,
        'No. detected trees': No_detected_trees_dominant,
        'DR (%)': dr_dominant,
        'No. correctly detected trees': matched_gt_ins_dominant,
        'DA (%)': da_dominant,
        'No. omitted trees': No_omitted_trees_dominant,
        'omission_error (%)': omission_error_dominant,
        'No. wrong detections': No_wrong_detected_trees_dominant,
        'commission_error (%)': commission_error_dominant,
        'F1_score (%)': F1_score_dominant,
        'mean_IoU (coverage)': mean_IoU_dominant
    }
    #Convert the dictionary to a list of dictionaries
    additional_metrics_list2 = [additional_metrics_data2]
    additional_metrics_df2 = pd.DataFrame(additional_metrics_list2)
        
    ############metrics for semantic segmentation#######################
    pred_sem_complete = pred_sem_complete+1
    gt_sem_complete = gt_sem_complete+1
    NUM_CLASSES_sem = 6
    sem_classcount_have = []
    sem_classcount = [1, 2, 3, 4, 5]
    sem_classcount_remove_ground = [1, 3, 4, 5]
    # acc and macc
    true_positive_classes = np.zeros(NUM_CLASSES_sem)
    positive_classes = np.zeros(NUM_CLASSES_sem)
    gt_classes = np.zeros(NUM_CLASSES_sem)
    # pn semantic mIoU
    for j in range(gt_sem_complete.shape[0]):
        gt_l = int(gt_sem_complete[j])
        pred_l = int(pred_sem_complete[j])
        gt_classes[gt_l] += 1
        positive_classes[pred_l] += 1
        true_positive_classes[gt_l] += int(gt_l == pred_l)
        
    # semantic results
    iou_list = []
    for i in range(NUM_CLASSES_sem):
        if gt_classes[i] > 0:
            sem_classcount_have.append(i)
            iou = true_positive_classes[i]/float(gt_classes[i]+positive_classes[i]-true_positive_classes[i]) 
        else:
            iou = 0.0
        iou_list.append(iou)
        
    set1 = set(sem_classcount)
    set2 = set(sem_classcount_have)
    set3 = set1 & set2
    sem_classcount_final = list(set3)

    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value='Semantic Segmentation oAcc')
    ws.cell(row=1, column=2, value=sum(true_positive_classes) / float(sum(positive_classes)))
    ws.cell(row=2, column=1, value='Semantic Segmentation mAcc')
    ws.cell(row=2, column=2, value=np.mean(np.array(true_positive_classes)[sem_classcount_final] / np.array(gt_classes)[sem_classcount_final]))

    row = 3
    ws.cell(row=3, column=1, value='Semantic Segmentation IoU')
    column = 2
    for iou in iou_list:
        ws.cell(row=row, column=column, value=iou)
        column += 1
        
    ws.cell(row=4, column=1, value='Semantic Segmentation mIoU')
    ws.cell(row=4, column=2, value=1. * sum(iou_list) / len(sem_classcount_final))
    
    ##remove ground points
    set1 = set(sem_classcount_remove_ground)
    set2 = set(sem_classcount_have)
    set3 = set1 & set2
    sem_classcount_final = list(set3)

    ws.cell(row=5, column=1, value='Semantic Segmentation oAcc without ground points')
    ws.cell(row=5, column=2, value=sum(true_positive_classes[sem_classcount_final]) / float(sum(positive_classes[[sem_classcount_final]])))
    ws.cell(row=6, column=1, value='Semantic Segmentation mAcc without ground points')
    ws.cell(row=6, column=2, value=np.mean(np.array(true_positive_classes)[sem_classcount_final] / np.array(gt_classes)[sem_classcount_final]))

    row = 7
    ws.cell(row=7, column=1, value='Semantic Segmentation IoU without ground points')
    column = 2
    iou_list = np.array(iou_list)
    for iou in iou_list[sem_classcount_final]:
        ws.cell(row=row, column=column, value=iou)
        column += 1
        
    ws.cell(row=8, column=1, value='Semantic Segmentation mIoU without ground points')
    ws.cell(row=8, column=2, value=1. * np.sum(iou_list[sem_classcount_final]) / len(sem_classcount_final))
    
    cm = confusion_matrix(gt_sem_complete, pred_sem_complete)
    ws.cell(row=9, column=1, value='Confusion Matrix:')
    for i in range(len(cm)):
        for j in range(len(cm[i])):
            ws.cell(row=10 + i, column=1 + j, value=str(cm[i][j]))
        
    tmp_output_path = current_plot+'/semantic_output.xlsx'
    wb.save(tmp_output_path)
    
    output_path = join(current_plot, 'instance_output.xlsx')
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df1.to_excel(writer, sheet_name='tree_instance_parameters', index=False)
        
        # Write the additional metrics DataFrame to the second sheet (individual tree detection)
        additional_metrics_df.to_excel(writer, sheet_name='individual tree detection', index=False)
        workbook = writer.book
        worksheet = writer.sheets['tree_instance_parameters']
        for i, col in enumerate(df1.columns):
            column_len = len(str(df1[col][0])) + 10
            worksheet.set_column(i, i, column_len)
        
        # Set the column widths for the second sheet (individual tree detection)
        worksheet2 = writer.sheets['individual tree detection']
        for i, col in enumerate(additional_metrics_df.columns):
            column_len = max(len(col), len(str(additional_metrics_df[col][0]))) + 2
            worksheet2.set_column(i, i, column_len)
            
        additional_metrics_df2.to_excel(writer, sheet_name='tree detection dominant', index=False)
        # Set the column widths for the second sheet (individual tree detection)
        worksheet3 = writer.sheets['tree detection dominant']
        for i, col in enumerate(additional_metrics_df2.columns):
            column_len = max(len(col), len(str(additional_metrics_df2[col][0]))) + 2
            worksheet3.set_column(i, i, column_len)
            
        writer.save()