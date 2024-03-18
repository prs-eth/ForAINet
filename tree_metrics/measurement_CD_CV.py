import os
import numpy as np
from scipy import stats
from utils.ply import read_ply, write_ply
from utils.utils import output_DTM_as_pc, DTM_generation, cal_DBH_and_centerP, DTM_accuracy, hdbscan_filtering, preprocess_point_cloud,alpha_shape_volume,plot_alpha_shape, compute_volume_with_voxelization, compute_volume_with_convex_hull
from os.path import exists, join
import matplotlib.pyplot as plt
import pyransac3d as pyrsc
from matplotlib.patches import Circle
import imageio
from skimage import color
from skimage import io
from RANSAC.RANSAC.RANSACCircle_2 import run
from plyfile import PlyData, PlyElement
from RANSAC.smallest_enclosing_circle import welzl
import math
from RANSAC.Common.Point import Point
from typing import List
from RANSAC.RANSAC.MatplotUtil import create_list_from_points
import pandas as pd
from scipy.interpolate import RectBivariateSpline, interp2d, bisplrep, bisplev, griddata
from scipy.spatial import ConvexHull
from openpyxl import Workbook
from sklearn.metrics import confusion_matrix
from tree_metrics.rmse import rmse
fold= ['/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/CULS/CULS_plot_2_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/NIBIO/NIBIO_plot_1_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/TUWIEN/TUWIEN_test_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/NIBIO/NIBIO_plot_17_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/NIBIO/NIBIO_plot_18_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/NIBIO/NIBIO_plot_22_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/NIBIO/NIBIO_plot_23_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/NIBIO/NIBIO_plot_5_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/RMIT/RMIT_test_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/SCION/SCION_plot_31_annotated_test.ply', '/scratch2/OutdoorPanopticSeg_V2/data_set1_5classes/treeinsfused/raw/SCION/SCION_plot_61_annotated_test.ply']
#Input files directory
file_path = '/scratch2/OutdoorPanopticSeg_V2/outputs/best_treemix2/eval/2023-08-26_21-51-41/'
all_files = os.listdir(file_path)
matching_files = [filename for filename in all_files if 'Semantic_results_forEval' in filename]
#output files directory
directory = file_path+'para_cal_imgs'
if not os.path.exists(directory):
    os.makedirs(directory)

#predefined parameters
BINSIZE = 0.5 #for dtm rasterization
height_of_trunkDiameter = 1.3
#th_height_of_tree = 5
thre_dbh = 0.1  #thresholf for dominant and non-dominant tree

#Iterate over all plots
for f_idx, filename in enumerate(matching_files):
    #predicted semantic segmentation file path
    pred_class_label_filename = file_path + filename
    #predicted instance segmentation file path
    pred_ins_label_filename = file_path +filename.replace('Semantic_results_forEval', 'Instance_Results_forEval')

    new_dir_name = "plot" + pred_class_label_filename[-6:-4]
    # Construct the full path to the new directory for new plot
    current_plot = os.path.join(directory, new_dir_name)
    # Check if the directory exists, and create it if it doesn't
    if os.path.exists(current_plot):
        continue
    if not os.path.exists(current_plot):
        os.mkdir(current_plot)

    #read files
    data_class = PlyData.read(pred_class_label_filename) 
    data_ins = PlyData.read(pred_ins_label_filename)
    pred_ins_complete = data_ins['vertex']['preds']
    pred_sem_complete = data_class['vertex']['preds']
    gt_ins_complete = data_ins['vertex']['gt'] -1
    gt_sem_complete = data_class['vertex']['gt']

    #idx for ground points
    idx_gt_ground = (gt_sem_complete==1)
    idx_pre_ground = (pred_sem_complete==1)

    ##########################Attributes for plot-level###########################
    #DTM generation from ground points
    #DTM of gt
    dtm_gt = DTM_generation(data_ins['vertex']['x'][idx_gt_ground], data_ins['vertex']['y'][idx_gt_ground], data_ins['vertex']['z'][idx_gt_ground], BINSIZE)
    #DTM of prediction
    dtm_pre = DTM_generation(data_ins['vertex']['x'][idx_pre_ground], data_ins['vertex']['y'][idx_pre_ground], data_ins['vertex']['z'][idx_pre_ground], BINSIZE)
    #output DTM as point cloud
    pc_out_dest = current_plot+'/floor_gt_pynn'+pred_class_label_filename[-6:-4]+'.ply'
    output_DTM_as_pc(dtm_gt, pc_out_dest)
    pc_out_dest = current_plot+'/floor_pre_pynn'+pred_class_label_filename[-6:-4]+'.ply'
    output_DTM_as_pc(dtm_pre, pc_out_dest)
    
    #The accuracy of the DTM
    dtm_coverage, dtm_RMSE = DTM_accuracy(data_ins['vertex'], BINSIZE, idx_gt_ground, idx_pre_ground)
    
    #Stem density for classify complexity of the plot
    #idx for vegetation points
    idx_gt_tree = (gt_sem_complete==2)|(gt_sem_complete==3)|(gt_sem_complete==4)
    idx_pre_tree = (pred_sem_complete==2)|(pred_sem_complete==3)|(pred_sem_complete==4)

    #GT instances for vegetation
    x_veg_gt = data_ins['vertex']['x'][idx_gt_tree]
    y_veg_gt = data_ins['vertex']['y'][idx_gt_tree]
    z_veg_gt = data_ins['vertex']['z'][idx_gt_tree]
    veg_ins_gt = gt_ins_complete[idx_gt_tree]
    veg_sem_gt = gt_sem_complete[idx_gt_tree]
    cor_veg_ins_pre = -1*np.ones_like(veg_ins_gt)
    veg_height_gt = np.zeros_like(veg_ins_gt, dtype=float)
    veg_Tdiamater_gt = np.zeros_like(veg_ins_gt, dtype=float)  #Trunk diameter
    veg_Cdiamater_gt = np.zeros_like(veg_ins_gt, dtype=float)  #Crown diameter
    veg_CV_gt = np.zeros_like(veg_ins_gt, dtype=float)  #Crown volume
    veg_CVl_gt = np.zeros_like(veg_ins_gt, dtype=float)  #Crown volume (live)
    veg_xloc_gt = np.zeros_like(veg_ins_gt, dtype=float)  #tree x location
    veg_yloc_gt = np.zeros_like(veg_ins_gt, dtype=float)  #tree y location

    #predicted instances for vegetation
    x_veg_pre = data_ins['vertex']['x'][idx_pre_tree]
    y_veg_pre = data_ins['vertex']['y'][idx_pre_tree]
    z_veg_pre = data_ins['vertex']['z'][idx_pre_tree]
    veg_ins_pre = pred_ins_complete[idx_pre_tree]
    veg_sem_pre = pred_sem_complete[idx_pre_tree]
    veg_height_pre = np.zeros_like(veg_ins_pre, dtype=float)
    veg_Tdiamater_pre =  np.zeros_like(veg_ins_pre, dtype=float)  #Trunk diameter
    veg_Cdiamater_pre =  np.zeros_like(veg_ins_pre, dtype=float)  #Crown diameter
    veg_CV_pre = np.zeros_like(veg_ins_pre, dtype=float)  #Crown volume
    veg_CVl_pre = np.zeros_like(veg_ins_pre, dtype=float)  #Crown volume (live)
    veg_xloc_pre = np.zeros_like(veg_ins_pre, dtype=float)  #tree x location
    veg_yloc_pre = np.zeros_like(veg_ins_pre, dtype=float)  #tree y location
    
    #tree instances in prediction set
    # initiate plot
    fig = plt.figure()
    #print(fig.dpi)  #dpi=100
    un = np.unique(veg_ins_pre)
    pts_in_pred = []
    pts_in_pred_tree = []
    #spline of ground
    interp_spline = bisplrep(dtm_pre[:,0], dtm_pre[:,1], dtm_pre[:,2])
    ##########################Attributes for tree-level###########################
    #loop for each predicted tree instance
    for ig, g in enumerate(un):
        fig.clf() # Clear old figure
        ax = fig.add_subplot(2, 4, 1, projection='3d')
        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        ax.set_zlabel('Z')
        if g == -1:
            continue
        tmp = np.where(veg_ins_pre == g)
        x_tmp = x_veg_pre[tmp]
        y_tmp = y_veg_pre[tmp]
        z_tmp = np.copy(z_veg_pre[tmp])
        
        #get ground height based on DTM
        x_center = np.mean(x_tmp)
        y_center = np.mean(y_tmp)
        ground_z = bisplev(x_center, y_center, interp_spline)
        #ground_z = griddata((dtm_pre[:,0], dtm_pre[:,1]), dtm_pre[:,2], (x_tmp, y_tmp), method='linear')
        #z value above ground
        z_tmp = z_tmp-ground_z
        
        tmp_tree = np.where(veg_ins_pre == g)
        tmp_com = np.where(pred_ins_complete == g)
        pts_in_pred += [tmp_com[0]]
        pts_in_pred_tree += [tmp_tree[0]]
        
        #show current tree instance to be calculated
        ax.scatter(x_tmp, y_tmp, z_tmp, marker='o', s=0.2)
        plt.pause(1)
        
        #hdbscan filtering
        xyz_tmp = np.concatenate((np.concatenate((np.expand_dims(x_tmp,axis=1),np.expand_dims(y_tmp,axis=1)),axis=1), np.expand_dims(z_tmp,axis=1)), axis=1)
        filtered_points, con_4 = hdbscan_filtering(xyz_tmp) #, distance_threshold=0.3)
        
        # Set the minimum points threshold
        min_points_threshold = 10
        im_path = current_plot+'/'+ 'preTree'+str(g)+ '_fittingPointsProj.png'
        
        if np.size(con_4)<min_points_threshold:
            fig1_path = im_path.replace('fittingPointsProj', '')
            plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=0.5, hspace=0.5)
            fig.savefig(fig1_path, bbox_inches='tight', pad_inches=0)
            continue
        veg_height_pre[tmp] = np.max(filtered_points[:,2])
         
        ax = fig.add_subplot(2, 4, 2, projection='3d')
        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        ax.set_zlabel('Z')
        #show tree instance after filtering
        ax.scatter(filtered_points[:,0], filtered_points[:,1], filtered_points[:,2], marker='o', s=0.2)
        plt.pause(1)
        
        ##############Calculate crown volume#################
        # Select points that belong to the live branches class (assumed to be labeled as 3)
        con_5 = np.where(veg_sem_pre[tmp] == 3)
        # Select points that belong to the branches class (assumed to be labeled as 3 or 4)
        con_6 = np.where((veg_sem_pre[tmp] == 3) | (veg_sem_pre[tmp] == 4))
        idx_for_fitting1 = np.intersect1d(con_4, con_5)
        points = xyz_tmp[idx_for_fitting1]
        
        idx_for_fitting2 = np.intersect1d(con_4, con_6)
        points2 = xyz_tmp[idx_for_fitting2]
                
        if np.size(idx_for_fitting2) < 10:
            fig1_path = im_path.replace('fittingPointsProj', '')
            plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=0.5, hspace=0.5)
            fig.savefig(fig1_path, bbox_inches='tight', pad_inches=0)
            continue
        
        #plot projected points and finding the smallest enclosing circle for tree crown
        xy_tmp = points2[:,0:2]
        # https://rosettacode.org/wiki/Smallest_enclosing_circle_problem#Python
        # Welzl's algorithm for smallest-circle problem
        # https://en.wikipedia.org/wiki/Smallest-circle_problem#Welzl's_algorithm
        nsphere = welzl(xy_tmp)
        #print("For points: ", xy_tmp)
        #print("Center is at: ", nsphere.center)
        #print("Radius of the smallest enclosing circle is: ", np.sqrt(nsphere.sqradius), "m\n")

        ax3 = fig.add_subplot(2, 4, 3)
        ax3.scatter(xy_tmp[:,0], xy_tmp[:,1], marker='.', s=0.1)
        ax3.scatter(nsphere.center[0], nsphere.center[1], marker='^')
        ax3.set_aspect(1)
        circle = plt.Circle((nsphere.center[0], nsphere.center[1]), np.sqrt(nsphere.sqradius), color='r', fill=False)
        ax3.add_patch(circle)
        ax3.set_title("crown radius = %2.3f m" % np.sqrt(nsphere.sqradius))

        # Calculate the Alpha Shape and its volume
        #volume = compute_volume_with_voxelization(points, im_path)
        #volume2 = compute_volume_with_voxelization(points2, im_path)
        
        if np.size(idx_for_fitting1) > 10:
            volume = compute_volume_with_convex_hull(points, im_path)
        else:
            volume = 0
        volume2 = compute_volume_with_convex_hull(points2, im_path)

        #plt.tight_layout()
        #plt.show()
        plt.pause(1)
        fig1_path = im_path.replace('fittingPointsProj', '')
        
        plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=0.5, hspace=0.5)
        fig.savefig(fig1_path, bbox_inches='tight', pad_inches=0)
        
        veg_Cdiamater_pre[tmp] = 2*np.sqrt(nsphere.sqradius) #Crown diameter
        veg_CV_pre[tmp] = volume2 
        veg_CVl_pre[tmp] = volume #Crown volume
    plt.close()  
    #tree instances in GT set
    fig = plt.figure()
    #print(fig.dpi)  #dpi=100
    un = np.unique(veg_ins_gt)
    pts_in_gt = []
    pts_in_gt_tree = []
    interp_spline = bisplrep(dtm_gt[:,0], dtm_gt[:,1], dtm_gt[:,2])
    for ig, g in enumerate(un):
        fig.clf() # Clear figure
        ax = fig.add_subplot(2, 4, 1, projection='3d')
        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        ax.set_zlabel('Z')
        if g == -1:
            continue
        tmp = np.where(veg_ins_gt == g)
        x_tmp = x_veg_gt[tmp]
        y_tmp = y_veg_gt[tmp]
        z_tmp = np.copy(z_veg_gt[tmp])
        
        #get ground height based on DTM
        x_center = np.mean(x_tmp)
        y_center = np.mean(y_tmp)
        ground_z = bisplev(x_center, y_center, interp_spline)
        #ground_z = griddata((dtm_pre[:,0], dtm_pre[:,1]), dtm_pre[:,2], (x_tmp, y_tmp), method='linear')
        #z value above ground
        z_tmp = z_tmp-ground_z
        
        tmp_tree = np.where(veg_ins_gt == g)
        tmp_com = np.where(gt_ins_complete == g) 
        pts_in_gt += [tmp_com[0]]
        pts_in_gt_tree += [tmp_tree[0]]
        
        #show current instance to be calculated
        ax.scatter(x_tmp, y_tmp, z_tmp, marker='o', s=0.2)
        plt.pause(1)
        
        #hdbscan filtering
        xyz_tmp = np.concatenate((np.concatenate((np.expand_dims(x_tmp,axis=1),np.expand_dims(y_tmp,axis=1)),axis=1), np.expand_dims(z_tmp,axis=1)), axis=1)
        filtered_points, con_4 = hdbscan_filtering(xyz_tmp)
        
        im_path = current_plot+'/'+ 'gtTree'+str(g)+ '_fittingPointsProj.png'
        if np.size(con_4)<min_points_threshold:
            fig1_path = im_path.replace('fittingPointsProj', '')
            plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=0.5, hspace=0.5)
            fig.savefig(fig1_path, bbox_inches='tight', pad_inches=0)
            continue
        veg_height_gt[tmp] = np.max(filtered_points[:,2])  #np.max(z_tmp)
        
        #xyz_tmp = np.concatenate((np.concatenate((np.expand_dims(x_tmp,axis=1),np.expand_dims(y_tmp,axis=1)),axis=1), np.expand_dims(z_tmp,axis=1)), axis=1)
        
        ax = fig.add_subplot(2, 4, 2, projection='3d')
        ax.set_xlabel('X')
        ax.set_ylabel('Y')
        ax.set_zlabel('Z')
        #show tree instance after filtering
        ax.scatter(filtered_points[:,0], filtered_points[:,1], filtered_points[:,2], marker='o', s=0.2)
        plt.pause(1)
        
        ##############Calculate crown volume and diameter#################
        # Select points that belong to the live branches class (assumed to be labeled as 3)
        con_5 = np.where(veg_sem_gt[tmp] == 3)
        # Select points that belong to the branches class (assumed to be labeled as 3 or 4)
        con_6 = np.where((veg_sem_gt[tmp] == 3) | (veg_sem_gt[tmp] == 4))
        idx_for_fitting1 = np.intersect1d(con_4, con_5)
        points = xyz_tmp[idx_for_fitting1]
        
        idx_for_fitting2 = np.intersect1d(con_4, con_6)
        points2 = xyz_tmp[idx_for_fitting2]
        
        if np.size(idx_for_fitting2) < 10:
            fig1_path = im_path.replace('fittingPointsProj', '')
            plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=0.5, hspace=0.5)
            fig.savefig(fig1_path, bbox_inches='tight', pad_inches=0)
            continue
        
        #plot projected points and finding the smallest enclosing circle for tree crown
        xy_tmp = points2[:,0:2]
        # https://rosettacode.org/wiki/Smallest_enclosing_circle_problem#Python
        # Welzl's algorithm for smallest-circle problem
        # https://en.wikipedia.org/wiki/Smallest-circle_problem#Welzl's_algorithm
        nsphere = welzl(xy_tmp)
        #print("For points: ", xy_tmp)
        #print("Center is at: ", nsphere.center)
        #print("Radius of the smallest enclosing circle is: ", np.sqrt(nsphere.sqradius), "m\n")
        
        ax3 = fig.add_subplot(2, 4, 3)
        ax3.scatter(xy_tmp[:,0], xy_tmp[:,1], marker='.', s=0.1)
        ax3.scatter(nsphere.center[0], nsphere.center[1], marker='^')
        ax3.set_aspect(1)
        circle = plt.Circle((nsphere.center[0], nsphere.center[1]), np.sqrt(nsphere.sqradius), color='r', fill=False)
        ax3.add_patch(circle)
        ax3.set_title("crown radius = %2.3f" % np.sqrt(nsphere.sqradius))

        # Calculate the Alpha Shape and its volume        
        #volume = compute_volume_with_voxelization(points, im_path)
        #volume2 = compute_volume_with_voxelization(points2, im_path)
        if np.size(idx_for_fitting1) > 10:
            volume = compute_volume_with_convex_hull(points, im_path)
        else:
            volume = 0
        volume2 = compute_volume_with_convex_hull(points2, im_path)
        
        plt.pause(1)
        fig1_path = im_path.replace('fittingPointsProj', '')
        
        plt.subplots_adjust(left=None, bottom=None, right=None, top=None, wspace=0.5, hspace=0.5)
        fig.savefig(fig1_path, bbox_inches='tight', pad_inches=0)
        
        veg_Cdiamater_gt[tmp] = 2*np.sqrt(nsphere.sqradius) #Crown diameter
        veg_CV_gt[tmp] = volume2
        veg_CVl_gt[tmp] = volume
    plt.close()   
    # Evaluate tree measurements with regard to GT
    cor_gt = []
    cor_pre = []
    gt_H = []
    gt_DBH = []
    error_H = []
    error_TD = []
    error_xloc = []
    error_yloc = []
    error_CV = []
    error_CVl = []
    gt_CV = []
    gt_CVl = []
    gt_CDia = []
    squared_error_CD = []
    total_gt_ins = len(pts_in_gt)  #No_reference_trees
    matched_gt_ins = 0  #No_correct_detected_trees
    No_detected_trees = len(pts_in_pred)
    at=0.5
    sum_IoU = 0
    # loop for each GT instance
    for ig, ins_gt in enumerate(pts_in_gt):
        ovmax = 0
        # find the best matched prediction
        for ip, ins_pred in enumerate(pts_in_pred):
            union = len(list(set(ins_pred).union(set(ins_gt)))) 
            intersect = len(list(set(ins_pred).intersection(set(ins_gt)))) 
            iou = float(intersect) / union 

            if iou > ovmax:
                ovmax = iou
                ipmax = ip

        if ovmax >= at:
            matched_gt_ins+=1
            cor_gt = pts_in_gt_tree[ig]
            cor_pre = pts_in_pred_tree[ipmax]
            cor_veg_ins_pre[cor_gt] = pred_ins_complete[pts_in_pred[ipmax][0]]
            error_H += [veg_height_pre[cor_pre][0]-veg_height_gt[cor_gt][0]]
            #[np.square(veg_height_pre[cor_pre][0]-veg_height_gt[cor_gt][0])]
            gt_H += [veg_height_gt[cor_gt][0]]
            error_TD += [veg_Tdiamater_pre[cor_pre][0]-veg_Tdiamater_gt[cor_gt][0]]
            #[np.square(veg_Tdiamater_pre[cor_pre][0]-veg_Tdiamater_gt[cor_gt][0])]
            gt_DBH += [veg_Tdiamater_gt[cor_gt][0]]
            squared_error_CD += [np.square(veg_Cdiamater_pre[cor_pre][0]-veg_Cdiamater_gt[cor_gt][0])]
            error_xloc += [veg_xloc_pre[cor_pre][0]-veg_xloc_gt[cor_gt][0]]
            error_yloc += [veg_yloc_pre[cor_pre][0]-veg_yloc_gt[cor_gt][0]]
            error_CV +=[veg_CV_pre[cor_pre][0]-veg_CV_gt[cor_gt][0]]
            error_CVl +=[veg_CVl_pre[cor_pre][0]-veg_CVl_gt[cor_gt][0]]
            gt_CV += [veg_CV_gt[cor_gt][0]]
            gt_CVl += [veg_CVl_gt[cor_gt][0]]
            gt_CDia += [veg_Cdiamater_gt[cor_gt][0]]
        sum_IoU += ovmax
    
    mean_IoU = (sum_IoU/total_gt_ins)*100
    No_omitted_trees = total_gt_ins-matched_gt_ins
    No_wrong_detected_trees = No_detected_trees-matched_gt_ins
    dr = (No_detected_trees/total_gt_ins)*100 
    da = (matched_gt_ins/total_gt_ins)*100  #completeness rate/tree detection accuracy
    omission_error = 100-da
    commission_error = (No_wrong_detected_trees/No_detected_trees)*100
    F1_score = (2*(100-commission_error)*da)/(100-commission_error+da)
    
    #RMSE, Root Mean Squared Error
    gt_h_hat = np.mean(gt_H)
    RMSE_H = np.sqrt(np.mean(np.square(error_H)))
    RMSE_H_100 = RMSE_H/gt_h_hat
    bias_H = np.mean(error_H)
    gt_dbh_hat = np.mean(gt_DBH)
    RMSE_TD = np.sqrt(np.mean(np.square(error_TD)))
    RMSE_TD_100 = RMSE_TD/gt_dbh_hat
    bias_TD = np.mean(error_TD)
    RMSE_CD = np.sqrt(np.mean(squared_error_CD))
    RMSE_locx = np.sqrt(np.mean(np.square(error_xloc)))
    RMSE_locy = np.sqrt(np.mean(np.square(error_yloc)))
    RMSE_locxy = (RMSE_locx+RMSE_locy)/2
    RMSE_CV = np.sqrt(np.mean(np.square(error_CV)))
    RMSE_CVl = np.sqrt(np.mean(np.square(error_CVl)))
    gt_CV_hat = np.mean(gt_CV)
    gt_CVl_hat = np.mean(gt_CVl)
    RMSE_CV_100 = RMSE_CV/gt_CV_hat
    RMSE_CVl_100 = RMSE_CVl/gt_CVl_hat
    gt_CDia_hat = np.mean(gt_CDia)
    RMSE_CD_100 = RMSE_CD/gt_CDia_hat
    print('true positive={}%, {} out of {} instances'.format((matched_gt_ins/total_gt_ins)*100, matched_gt_ins, total_gt_ins))
    print('RMSE for vegetation height: {}'.format(RMSE_H))
    print('RMSE for vegetation trunk diameter: {}'.format(RMSE_TD))
    print('RMSE for vegetation crown diameter: {}'.format(RMSE_CD))

    # Project the point of trees onto the xy plane
    projected_points = np.concatenate((np.expand_dims(x_veg_gt,axis=1),np.expand_dims(y_veg_gt,axis=1)),axis=1)
    # Calculate the convex hull of the projected points
    hull = ConvexHull(projected_points)
    # Calculate the area of the convex hull polygon
    convex_area = hull.area
    stem_density = (total_gt_ins/convex_area*10000)  #square meters to hectares, 1 hectare (ha) = 10,000 square meters (m²)
    
    output_path = join(current_plot, 'vegetable_gt_total.ply')
    write_ply(output_path,
            [x_veg_gt, y_veg_gt, z_veg_gt, veg_ins_gt, veg_height_gt, veg_Tdiamater_gt, veg_Cdiamater_gt, cor_veg_ins_pre],
            ['x', 'y', 'z', 'ins_label', 'height_gt', 'trunk_diamater_gt', 'crown_diamater_gt', 'pre_label'])
    output_path2 = join(current_plot, 'vegetable_pre_total.ply')
    write_ply(output_path2,
            [x_veg_pre, y_veg_pre, z_veg_pre, veg_ins_pre, veg_height_pre, veg_Tdiamater_pre, veg_Cdiamater_pre],
            ['x', 'y', 'z', 'ins_label', 'height_pre', 'trunk_diamater_pre', 'crown_diamater_pre'])

    veg_ins_gt_u, indices = np.unique(veg_ins_gt, return_index=True)
    cor_pre_ins = [] #-1*np.ones_like(indices)
    cor_veg_height_pre = []
    cor_veg_Tdiamater_pre = []
    cor_veg_Cdiamater_pre = []
    cor_veg_CV_pre = []
    cor_veg_CVl_pre = []
    for i in cor_veg_ins_pre[indices]:
        if i==-1:
            #idx_for_pre.append(0)
            cor_pre_ins.append(-1)
            cor_veg_height_pre.append(-1)
            cor_veg_Tdiamater_pre.append(-1)
            cor_veg_Cdiamater_pre.append(-1)
            cor_veg_CV_pre.append(-1)
            cor_veg_CVl_pre.append(-1)
        else:
            inn=np.where(veg_ins_pre==i)[0][0] 
            cor_pre_ins.append(veg_ins_pre[inn])
            cor_veg_height_pre.append(veg_height_pre[inn])
            cor_veg_Tdiamater_pre.append(veg_Tdiamater_pre[inn])
            cor_veg_Cdiamater_pre.append(veg_Cdiamater_pre[inn])
            cor_veg_CV_pre.append(veg_CV_pre[inn])
            cor_veg_CVl_pre.append(veg_CVl_pre[inn])
    cor_pre_ins = np.array(cor_pre_ins)
    cor_veg_height_pre = np.array(cor_veg_height_pre)
    cor_veg_Tdiamater_pre = np.array(cor_veg_Tdiamater_pre)
    cor_veg_Cdiamater_pre = np.array(cor_veg_Cdiamater_pre)
    cor_veg_CV_pre = np.array(cor_veg_CV_pre)
    cor_veg_CVl_pre = np.array(cor_veg_CVl_pre)
    c_ind = np.where(veg_height_gt[indices]!=0)

    #get field DBH data
    gt_labels = veg_ins_gt[indices[c_ind]]
    current_test_file_name = fold[f_idx].split('/')[-1]
    field_file_path = '/scratch2/tree_metrics/field_gt/'
    file_to_find = current_test_file_name.split('_')[0]
    file_to_find = f"tree_data_{file_to_find}.csv"
    file_full_path = os.path.join(field_file_path, file_to_find)
    plotID = None
    DBH_values = []
    for part in current_test_file_name.split('_'):
        if part.isdigit():
            plotID = int(part)
            break
    if os.path.exists(file_full_path):
        df = pd.read_csv(file_full_path)
        for gt_label in(gt_labels):
            DBH_value = -1
            if 'plotID' in df.columns:
                if plotID is not None:
                    selected_row = df[(df['plotID'] == plotID) & (df['treeID'] == gt_label)]
                else:
                    selected_row = df[df['treeID'] == gt_label]
            elif 'treeID' in df.columns:
                selected_row = df[df['treeID'] == gt_label]

            if not selected_row.empty:
                DBH_value = selected_row['DBH'].values[0]
            DBH_values.append(DBH_value)
    DBH_values = np.array(DBH_values)
    
    idx_valid = (DBH_values!=-1)
    if DBH_values[idx_valid][0]>1:
        DBH_values[idx_valid] = DBH_values[idx_valid]/100
    
    output = np.hstack((veg_ins_gt[indices[c_ind]].reshape(-1,1),cor_pre_ins[c_ind].reshape(-1,1),
            veg_height_gt[indices[c_ind]].reshape(-1,1), cor_veg_height_pre[c_ind].reshape(-1,1),
            veg_Tdiamater_gt[indices[c_ind]].reshape(-1,1), cor_veg_Tdiamater_pre[c_ind].reshape(-1,1),
            DBH_values.reshape(-1,1), 
            veg_Cdiamater_gt[indices[c_ind]].reshape(-1,1), cor_veg_Cdiamater_pre[c_ind].reshape(-1,1),
            veg_CV_gt[indices[c_ind]].reshape(-1,1), cor_veg_CV_pre[c_ind].reshape(-1,1),
            veg_CVl_gt[indices[c_ind]].reshape(-1,1), cor_veg_CVl_pre[c_ind].reshape(-1,1)))
    df1 = pd.DataFrame(output,
                    dtype=np.float32,
                    columns=['gt label', 'pre label',
                                'gt height', 'pre height',
                                'gt DBH', 'pre DBH', 'field DBH',
                                'gt crown diameter', 'pre crown diameter',
                                'gt crown volume', 'pre crown volume',
                                'gt crown volume (live)', 'pre crown volume (live)'])
    idx_valid = (DBH_values!=-1) & (cor_veg_Tdiamater_pre[c_ind]!=-1)
    
    RMSE_TD_field = rmse(DBH_values[idx_valid], cor_veg_Tdiamater_pre[c_ind][idx_valid])
    # Combine the additional metrics into a new DataFrame
    additional_metrics_data = {
        'dtm_coverage (%)' : dtm_coverage,
        'dtm_RMSE' : dtm_RMSE,
        'No. reference trees': total_gt_ins,
        'No. detected trees': No_detected_trees,
        'DR (%)': dr,
        'No. correctly detected trees': matched_gt_ins,
        'DA (%)': da,
        'No. omitted trees': No_omitted_trees,
        'omission_error (%)': omission_error,
        'No. wrong detections': No_wrong_detected_trees,
        'commission_error (%)': commission_error,
        'F1_score (%)': F1_score,
        'mean_IoU (coverage)': mean_IoU,
        'stem_density': stem_density,
        'RMSE for tree height': RMSE_H,
        'RMSE% for tree height': RMSE_H_100,
        'bias for tree height': bias_H,
        'RMSE for DBH': RMSE_TD,
        'RMSE% for DBH': RMSE_TD_100,
        'bias for DBH': bias_TD,
        'RMSE for field DBH': RMSE_TD_field,
        'RMSE for crown diameter': RMSE_CD,
        'RMSE% for crown diameter': RMSE_CD_100,
        'RMSE for location x': RMSE_locx,
        'RMSE for location y': RMSE_locy,
        'RMSE for location': RMSE_locxy,
        'RMSE for crown volume': RMSE_CV,
        'RMSE for crown volume (live)': RMSE_CVl,
        'RMSE% for crown volume': RMSE_CV_100,
        'RMSE% for crown volume (live)': RMSE_CVl_100
    }
    #Convert the dictionary to a list of dictionaries
    additional_metrics_list = [additional_metrics_data]
    additional_metrics_df = pd.DataFrame(additional_metrics_list)
        
    ############metrics for semantic segmentation#######################
    pred_sem_complete = pred_sem_complete+1
    gt_sem_complete = gt_sem_complete+1
    NUM_CLASSES_sem = 6
    sem_classcount_have = []
    sem_classcount = [1, 2, 3, 4, 5]
    sem_classcount_remove_ground = [1, 3, 4, 5]
    # acc and macc
    true_positive_classes = np.zeros(NUM_CLASSES_sem)
    positive_classes = np.zeros(NUM_CLASSES_sem)
    gt_classes = np.zeros(NUM_CLASSES_sem)
    # pn semantic mIoU
    for j in range(gt_sem_complete.shape[0]):
        gt_l = int(gt_sem_complete[j])
        pred_l = int(pred_sem_complete[j])
        gt_classes[gt_l] += 1
        positive_classes[pred_l] += 1
        true_positive_classes[gt_l] += int(gt_l == pred_l)
        
    # semantic results
    iou_list = []
    for i in range(NUM_CLASSES_sem):
        if gt_classes[i] > 0:
            sem_classcount_have.append(i)
            iou = true_positive_classes[i]/float(gt_classes[i]+positive_classes[i]-true_positive_classes[i]) 
        else:
            iou = 0.0
        iou_list.append(iou)
        
    set1 = set(sem_classcount)
    set2 = set(sem_classcount_have)
    set3 = set1 & set2
    sem_classcount_final = list(set3)

    wb = Workbook()
    ws = wb.active

    ws.cell(row=1, column=1, value='Semantic Segmentation oAcc')
    ws.cell(row=1, column=2, value=sum(true_positive_classes) / float(sum(positive_classes)))
    ws.cell(row=2, column=1, value='Semantic Segmentation mAcc')
    ws.cell(row=2, column=2, value=np.mean(np.array(true_positive_classes)[sem_classcount_final] / np.array(gt_classes)[sem_classcount_final]))

    row = 3
    ws.cell(row=3, column=1, value='Semantic Segmentation IoU')
    column = 2
    for iou in iou_list:
        ws.cell(row=row, column=column, value=iou)
        column += 1
        
    ws.cell(row=4, column=1, value='Semantic Segmentation mIoU')
    ws.cell(row=4, column=2, value=1. * sum(iou_list) / len(sem_classcount_final))
    
    ##remove ground points
    set1 = set(sem_classcount_remove_ground)
    set2 = set(sem_classcount_have)
    set3 = set1 & set2
    sem_classcount_final = list(set3)

    ws.cell(row=5, column=1, value='Semantic Segmentation oAcc without ground points')
    ws.cell(row=5, column=2, value=sum(true_positive_classes[sem_classcount_final]) / float(sum(positive_classes[[sem_classcount_final]])))
    ws.cell(row=6, column=1, value='Semantic Segmentation mAcc without ground points')
    ws.cell(row=6, column=2, value=np.mean(np.array(true_positive_classes)[sem_classcount_final] / np.array(gt_classes)[sem_classcount_final]))

    row = 7
    ws.cell(row=7, column=1, value='Semantic Segmentation IoU without ground points')
    column = 2
    iou_list = np.array(iou_list)
    for iou in iou_list[sem_classcount_final]:
        ws.cell(row=row, column=column, value=iou)
        column += 1
        
    ws.cell(row=8, column=1, value='Semantic Segmentation mIoU without ground points')
    ws.cell(row=8, column=2, value=1. * np.sum(iou_list[sem_classcount_final]) / len(sem_classcount_final))
    
    cm = confusion_matrix(gt_sem_complete, pred_sem_complete)
    ws.cell(row=9, column=1, value='Confusion Matrix:')
    for i in range(len(cm)):
        for j in range(len(cm[i])):
            ws.cell(row=10 + i, column=1 + j, value=str(cm[i][j]))
        
    tmp_output_path = current_plot+'/semantic_output.xlsx'
    wb.save(tmp_output_path)
    
    output_path = join(current_plot, 'instance_output.xlsx')
    with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
        df1.to_excel(writer, sheet_name='tree_instance_parameters', index=False)
        
        # Write the additional metrics DataFrame to the second sheet (individual tree detection)
        additional_metrics_df.to_excel(writer, sheet_name='individual tree detection', index=False)
        workbook = writer.book
        worksheet = writer.sheets['tree_instance_parameters']
        for i, col in enumerate(df1.columns):
            column_len = len(str(df1[col][0])) + 10
            worksheet.set_column(i, i, column_len)
        
        # Set the column widths for the second sheet (individual tree detection)
        worksheet2 = writer.sheets['individual tree detection']
        for i, col in enumerate(additional_metrics_df.columns):
            column_len = max(len(col), len(str(additional_metrics_df[col][0]))) + 2
            worksheet2.set_column(i, i, column_len)
            
        #writer.book = wb
        #writer.sheets['semantic segmentation'] = ws
        writer.save()